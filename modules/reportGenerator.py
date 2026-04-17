"""
modules/reportGenerator.py
Generates a fully-detailed, professionally styled Excel report.

Sheets:
  1. Summary          — KPIs + pie chart
  2. File Overview    — one row per file, ALL columns
  3. Issues Detail    — one row per issue with line number + line content + suggestion
  4. Skipped & Failed — dedicated sheet for incomplete scans

Every column required by spec is present and populated.
"""

import os
from typing import Dict, Any, List
from datetime import datetime

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import PieChart, BarChart, Reference
from openpyxl.chart.label import DataLabel

from modules.logger import getLogger

logger = getLogger()

# ── Palette ────────────────────────────────────────────────────────────────
C = {
    "hdrBg"   : "1A1A2E",   # dark navy header bg
    "hdrFg"   : "FFFFFF",   # white header text
    "rowEven" : "EEF2FF",   # light blue-grey alternating row
    "rowOdd"  : "FFFFFF",
    "high"    : "C0392B",   # red
    "highFg"  : "FFFFFF",
    "medium"  : "E67E22",   # orange
    "medFg"   : "FFFFFF",
    "low"     : "F1C40F",   # yellow
    "lowFg"   : "000000",
    "clean"   : "27AE60",   # green
    "cleanFg" : "FFFFFF",
    "skipped" : "7F8C8D",   # grey
    "skipFg"  : "FFFFFF",
    "failed"  : "922B21",   # dark red
    "failFg"  : "FFFFFF",
    "success" : "2471A3",   # blue
    "succFg"  : "FFFFFF",
    "frontend": "1A5276",   # dark blue
    "backend" : "145A32",   # dark green
    "config"  : "6E2F8B",   # purple
    "binary"  : "555555",   # dark grey
    "unknown" : "7D6608",   # dark yellow
    "kpiBg"   : "F8F9FA",
    "kpiBdr"  : "DEE2E6",
}

import re

def cleanText(text):
    if not isinstance(text, str):
        return text

    # Remove illegal characters
    text = re.sub(r'[\x00-\x08\x0B-\x0C\x0E-\x1F]', '', text)

    # Remove invalid unicode
    text = text.encode("utf-8", "ignore").decode("utf-8")

    return text
# ══════════════════════════════════════════════════════════════════════════
# Public entry point
# ══════════════════════════════════════════════════════════════════════════

def generateExcelReport(scanResults: Dict[str, Any], outputPath: str = None) -> str:
    """
    Build and save the Excel workbook.
    Returns the absolute path of the saved file.
    """
    if not outputPath:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        outputPath = os.path.join("uploads", f"scan_report_{ts}.xlsx")

    dirName = os.path.dirname(outputPath)
    if dirName:
        os.makedirs(dirName, exist_ok=True)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)   # remove default blank sheet

    wsSummary  = wb.create_sheet("Summary")
    wsOverview = wb.create_sheet("File Overview")
    wsIssues   = wb.create_sheet("Issues Detail")
    wsSkipped  = wb.create_sheet("Skipped & Failed")

    files   = scanResults.get("files", [])
    summary = scanResults.get("summary", {})

    _buildSummarySheet(wsSummary,  summary, scanResults.get("repoUrl", ""))
    _buildOverviewSheet(wsOverview, files)
    _buildIssuesSheet(wsIssues,   files)
    _buildSkippedSheet(wsSkipped,  files)
    _addPieChart(wsSummary, summary)

    # ✅ CLEAN ALL SHEETS BEFORE SAVING (CRITICAL FIX)

    for sheet in wb.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str):
                    cell.value = cleanText(cell.value)

    wb.save(outputPath)
    logger.info(f"Excel report saved → {outputPath}")
    return os.path.abspath(outputPath)


# ══════════════════════════════════════════════════════════════════════════
# Sheet 1 — Summary
# ══════════════════════════════════════════════════════════════════════════

def _buildSummarySheet(ws, summary: Dict[str, Any], repoUrl: str):
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 4
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 22

    # ── Title ──
    ws.merge_cells("A1:E1")
    _cell(ws, "A1",
          "🔍  GitHub Repository Validator — Security & Quality Report",
          bold=True, size=15, fgText=C["hdrFg"], bgColor=C["hdrBg"],
          align="center", height=42)

    # ── Sub-title ──
    ws.merge_cells("A2:E2")
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    _cell(ws, "A2",
          f"Repository: {repoUrl or 'N/A'}   |   Generated: {ts}   |   Scan time: {summary.get('totalTime', 0)}s",
          size=10, italic=True, fgText="555555", bgColor="F8F9FA",
          align="center", height=22)

    ws.append([])  # spacer row 3

    # ── KPI cards (rows 4–12) ──
    sev = summary.get("severityCounts", {})
    kpis = [
        ("📁  Total Files Scanned",  summary.get("totalFiles",   0), "0070C0"),
        ("⚠️  Total Issues Found",   summary.get("totalIssues",  0), "C0392B" if summary.get("totalIssues") else "27AE60"),
        ("✅  Clean Files",          summary.get("cleanFiles",   0), "27AE60"),
        ("🔍  Files With Issues",    summary.get("successFiles", 0), "0070C0"),
        ("❌  Failed Files",         summary.get("failedFiles",  0), "C0392B"),
        ("⏭️  Skipped Files",        summary.get("skippedFiles", 0), "7F8C8D"),
        ("🔴  High Severity",        sev.get("High",   0),           "C0392B"),
        ("🟠  Medium Severity",      sev.get("Medium", 0),           "E67E22"),
        ("🟡  Low Severity",         sev.get("Low",    0),           "F1C40F"),
    ]

    for i, (label, value, colour) in enumerate(kpis, start=4):
        ws.row_dimensions[i].height = 26
        lc = ws.cell(row=i, column=1, value=label)
        vc = ws.cell(row=i, column=2, value=value)
        for cell in [lc, vc]:
            cell.border  = _border()
            cell.alignment = Alignment(vertical="center", indent=1)
        lc.font  = Font(name="Calibri", bold=True, size=11)
        lc.fill  = PatternFill("solid", fgColor=C["kpiBg"])
        vc.font  = Font(name="Calibri", bold=True, size=16, color=colour)
        vc.fill  = PatternFill("solid", fgColor="FAFAFA")
        vc.alignment = Alignment(horizontal="center", vertical="center")

    # ── Chart data (hidden — used by pie chart, cols D/E rows 4–7) ──
    chartLabels = ["Clean", "Has Issues", "Failed", "Skipped"]
    chartValues = [
        summary.get("cleanFiles",   0),
        summary.get("successFiles", 0),
        summary.get("failedFiles",  0),
        summary.get("skippedFiles", 0),
    ]
    for i, (lbl, val) in enumerate(zip(chartLabels, chartValues), start=4):
        ws.cell(row=i, column=4, value=lbl).font = Font(size=9, color="AAAAAA")
        ws.cell(row=i, column=5, value=val).font  = Font(size=9, color="AAAAAA")


# ══════════════════════════════════════════════════════════════════════════
# Sheet 2 — File Overview  (one row per file, ALL spec columns)
# ══════════════════════════════════════════════════════════════════════════

OV_HEADERS = [
    "#",
    "File Name",
    "File Path",
    "File Type\n(Extension)",
    "Code Type\n(Frontend/Backend/Config)",
    "Scan Status",
    "Issues\nFound",
    "Highest\nSeverity",
    "Scan Mode\n(Fast/Deep)",
    "Lines\nScanned",
    "Processing\nTime (s)",
    "LLM Deep\nScan Used",
    "File Truncated\n(>Size Limit)",
    "Error Message",
    "Skipped Reason\n(if any)",
]

OV_WIDTHS = [5, 28, 58, 13, 26, 14, 10, 12, 12, 10, 14, 14, 16, 45, 50]


def _buildOverviewSheet(ws, files: List[Dict[str, Any]]):
    ws.sheet_view.showGridLines = False
    _writeHeaderRow(ws, OV_HEADERS, OV_WIDTHS)

    for rowIdx, fr in enumerate(files, start=2):
        status   = fr.get("scanStatus",  "Unknown")
        codeType = fr.get("codeType",    "Unknown")
        issues   = fr.get("issues",      [])
        highestSev = _highestSeverity(issues)

        rowData = [
            rowIdx - 1,
            fr.get("fileName",       ""),
            fr.get("filePath",       ""),
            fr.get("fileType",       ""),
            codeType,
            status,
            fr.get("issuesFound",    0),
            highestSev if issues else "None",
            fr.get("scanMode",       "Fast"),
            fr.get("linesScanned",   fr.get("lineCount", 0)),
            fr.get("processingTime", 0.0),
            "Yes" if fr.get("scanMode") == "Deep" else "No",
            "Yes" if fr.get("truncated") else "No",
            fr.get("errorMessage",   ""),
            fr.get("skippedReason",  ""),
        ]

        bgCol = C["rowEven"] if rowIdx % 2 == 0 else C["rowOdd"]
        for colIdx, value in enumerate(rowData, start=1):
            try:
                clean_value = cleanText(value)   # 🔥 CLEAN HERE
                cell = ws.cell(row=rowIdx, column=colIdx, value=clean_value)
            except Exception:
                cell = ws.cell(row=rowIdx, column=colIdx, value="INVALID_DATA")
            cell.font      = Font(name="Calibri", size=10)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border    = _border()
            cell.fill      = PatternFill("solid", fgColor=bgCol)

        # Colour: Scan Status (col 6)
        _colourBadge(ws.cell(row=rowIdx, column=6),  status,   _statusColour)
        # Colour: Code Type (col 5)
        _colourBadge(ws.cell(row=rowIdx, column=5),  codeType, _codeTypeColour)
        # Colour: Highest Severity (col 8)
        _colourBadge(ws.cell(row=rowIdx, column=8),  highestSev if issues else "None", _severityColour)

        ws.row_dimensions[rowIdx].height = 20

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes    = "B2"


# ══════════════════════════════════════════════════════════════════════════
# Sheet 3 — Issues Detail  (one row per issue)
# ══════════════════════════════════════════════════════════════════════════

IS_HEADERS = [
    "#",
    "File Name",
    "File Path",
    "File Type",
    "Code Type\n(Frontend/Backend)",
    "Scan Mode",
    "Issue Type",
    "Description",
    "Line\nNumber",
    "Actual Line Content\n(Source Code)",
    "Severity",
    "Suggestion / Fix\n(Remediation)",
    "Validation\nType",
]

IS_WIDTHS = [5, 26, 52, 12, 22, 10, 22, 60, 9, 65, 10, 70, 16]


def _buildIssuesSheet(ws, files: List[Dict[str, Any]]):
    ws.sheet_view.showGridLines = False
    _writeHeaderRow(ws, IS_HEADERS, IS_WIDTHS)

    rowIdx   = 2
    issueNum = 1

    for fr in files:
        for issue in fr.get("issues", []):
            severity   = issue.get("severity",       "Low")
            lineNum    = issue.get("lineNumber",      0)
            lineContent = issue.get("lineContent",   "")
            suggestion  = issue.get("suggestion",    "")
            actualIssue = issue.get("actualIssue", issue.get("description", ""))

            # Ensure suggestion is never blank
            if not suggestion.strip():
                suggestion = _fallbackSuggestion(issue.get("issueType", ""), description)

            rowData = [
                issueNum,
                fr.get("fileName",  ""),
                fr.get("filePath",  ""),
                fr.get("fileType",  ""),
                fr.get("codeType",  ""),
                fr.get("scanMode",  "Fast"),
                issue.get("issueType",      ""),
                actualIssue,   # 🔥 SHOW REAL ISSUE
                lineNum if lineNum else "N/A",
                lineContent if lineContent else "(see description)",
                severity,
                suggestion,
                issue.get("validationType", ""),
            ]

            bgCol = C["rowEven"] if rowIdx % 2 == 0 else C["rowOdd"]
            for colIdx, value in enumerate(rowData, start=1):
                try:
                    clean_value = cleanText(value)   # 🔥 CLEAN HERE (VERY IMPORTANT)
                    cell = ws.cell(row=rowIdx, column=colIdx, value=clean_value)
                except Exception:
                    cell = ws.cell(row=rowIdx, column=colIdx, value=value)
                cell.font      = Font(name="Calibri", size=10)
                cell.alignment = Alignment(vertical="center", wrap_text=True)
                cell.border    = _border()
                cell.fill      = PatternFill("solid", fgColor=bgCol)

            # Colour severity badge (col 11)
            _colourBadge(ws.cell(row=rowIdx, column=11), severity, _severityColour)
            # Colour code type (col 5)
            _colourBadge(ws.cell(row=rowIdx, column=5),  fr.get("codeType", ""), _codeTypeColour)

            # Highlight line content cell in light yellow so it stands out
            ws.cell(row=rowIdx, column=10).fill = PatternFill("solid", fgColor="FFFDE7")
            ws.cell(row=rowIdx, column=10).font = Font(name="Courier New", size=9)

            ws.row_dimensions[rowIdx].height = 36
            rowIdx   += 1
            issueNum += 1

    if rowIdx == 2:
        # No issues found — write a friendly message
        ws.merge_cells("A2:M2")
        _cell(ws, "A2", "✅  No issues found — all scanned files are clean.",
              bold=True, size=12, fgText="27AE60", bgColor="EAFAF1", align="center", height=36)

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes    = "B2"


# ══════════════════════════════════════════════════════════════════════════
# Sheet 4 — Skipped & Failed
# ══════════════════════════════════════════════════════════════════════════

SK_HEADERS = [
    "#",
    "File Name",
    "File Path",
    "File Type",
    "Code Type",
    "Status",
    "Skipped Reason",
    "Error Message",
    "Lines Scanned\nBefore Skip",
    "Partial Issues\nFound",
    "Processing\nTime (s)",
    "Recommended Action",
]

SK_WIDTHS = [5, 28, 58, 12, 16, 12, 55, 55, 16, 16, 14, 60]


def _buildSkippedSheet(ws, files: List[Dict[str, Any]]):
    ws.sheet_view.showGridLines = False
    _writeHeaderRow(ws, SK_HEADERS, SK_WIDTHS)

    skippedFiles = [
        fr for fr in files
        if fr.get("scanStatus") in ("Skipped", "Failed")
    ]

    if not skippedFiles:
        ws.merge_cells("A2:L2")
        _cell(ws, "A2", "✅  No files were skipped or failed during this scan.",
              bold=True, size=11, fgText="27AE60", bgColor="EAFAF1", align="center", height=36)
        return

    for rowIdx, fr in enumerate(skippedFiles, start=2):
        status = fr.get("scanStatus", "Skipped")
        reason = fr.get("skippedReason", "")
        error  = fr.get("errorMessage", "")

        # Build recommended action from the reason
        action = _skippedAction(reason, error, fr.get("fileType", ""))

        rowData = [
            rowIdx - 1,
            fr.get("fileName",       ""),
            fr.get("filePath",       ""),
            fr.get("fileType",       ""),
            fr.get("codeType",       ""),
            status,
            reason  or "—",
            error   or "—",
            fr.get("linesScanned",   fr.get("lineCount", 0)),
            fr.get("issuesFound",    0),
            fr.get("processingTime", 0.0),
            action,
        ]

        bgCol = C["rowEven"] if rowIdx % 2 == 0 else C["rowOdd"]
        for colIdx, value in enumerate(rowData, start=1):
            try:
                clean_value = cleanText(value)   # ✅ CLEAN HERE
                cell = ws.cell(row=rowIdx, column=colIdx, value=clean_value)
            except Exception:
                cell = ws.cell(row=rowIdx, column=colIdx, value="INVALID_DATA")    
            cell.font      = Font(name="Calibri", size=10)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border    = _border()
            cell.fill      = PatternFill("solid", fgColor=bgCol)

        _colourBadge(ws.cell(row=rowIdx, column=6), status, _statusColour)
        ws.row_dimensions[rowIdx].height = 30

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes    = "B2"


# ══════════════════════════════════════════════════════════════════════════
# Chart
# ══════════════════════════════════════════════════════════════════════════

def _addPieChart(ws, summary: Dict[str, Any]):
    try:
        pie          = PieChart()
        pie.title    = "Scan Results Distribution"
        pie.style    = 26
        pie.width    = 16
        pie.height   = 12

        dataRef = Reference(ws, min_col=5, min_row=4, max_row=7)
        catRef  = Reference(ws, min_col=4, min_row=4, max_row=7)
        pie.add_data(dataRef)
        pie.set_categories(catRef)
        pie.dataLabels           = DataLabel()
        pie.dataLabels.showPercent   = True
        pie.dataLabels.showCatName   = True

        ws.add_chart(pie, "D4")
    except Exception as e:
        logger.warning(f"Pie chart skipped: {e}")


# ══════════════════════════════════════════════════════════════════════════
# Helpers
# ══════════════════════════════════════════════════════════════════════════

def _writeHeaderRow(ws, headers: List[str], widths: List[int]):
    for colIdx, (header, width) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(row=1, column=colIdx, value=header)
        cell.font      = Font(name="Calibri", bold=True, size=11, color=C["hdrFg"])
        cell.fill      = PatternFill("solid", fgColor=C["hdrBg"])
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = _border()
        ws.column_dimensions[get_column_letter(colIdx)].width = width
    ws.row_dimensions[1].height = 36


def _cell(ws, addr: str, value, bold=False, size=11, italic=False,
          fgText="000000", bgColor=None, align="left", height=None):
    cell = ws[addr]
    cell.value     = value
    cell.font      = Font(name="Calibri", bold=bold, size=size, italic=italic, color=fgText)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    if bgColor:
        cell.fill = PatternFill("solid", fgColor=bgColor)
    row = int("".join(filter(str.isdigit, addr)))
    if height:
        ws.row_dimensions[row].height = height


def _border() -> Border:
    s = Side(style="thin", color="D0D0D0")
    return Border(left=s, right=s, top=s, bottom=s)


def _colourBadge(cell, value: str, colourFn):
    bg, fg = colourFn(value)
    cell.fill      = PatternFill("solid", fgColor=bg)
    cell.font      = Font(name="Calibri", bold=True, size=10, color=fg)
    cell.alignment = Alignment(horizontal="center", vertical="center")


def _statusColour(status: str):
    return {
        "Clean"  : (C["clean"],   C["cleanFg"]),
        "Success": (C["success"], C["succFg"]),
        "Failed" : (C["failed"],  C["failFg"]),
        "Skipped": (C["skipped"], C["skipFg"]),
        "None"   : ("CCCCCC",     "333333"),
    }.get(status, ("CCCCCC", "333333"))


def _severityColour(sev: str):
    return {
        "High"  : (C["high"],    C["highFg"]),
        "Medium": (C["medium"],  C["medFg"]),
        "Low"   : (C["low"],     C["lowFg"]),
        "None"  : (C["clean"],   C["cleanFg"]),
    }.get(sev, ("CCCCCC", "333333"))


def _codeTypeColour(ctype: str):
    return {
        "Frontend": (C["frontend"], "FFFFFF"),
        "Backend" : (C["backend"],  "FFFFFF"),
        "Config"  : (C["config"],   "FFFFFF"),
        "Binary"  : (C["binary"],   "FFFFFF"),
        "Unknown" : (C["unknown"],  "FFFFFF"),
    }.get(ctype, ("AAAAAA", "FFFFFF"))


def _highestSeverity(issues: List[Dict[str, Any]]) -> str:
    order = {"High": 3, "Medium": 2, "Low": 1}
    best  = 0
    label = "Low"
    for issue in issues:
        sev = issue.get("severity", "Low")
        if order.get(sev, 0) > best:
            best  = order[sev]
            label = sev
    return label


def _skippedAction(reason: str, error: str, fileType: str) -> str:
    combined = (reason + " " + error).lower()
    if "binary" in combined:
        return "Binary files do not require code scanning. No action needed."
    if "permission" in combined:
        return "Fix file permissions so the scanner can read this file: chmod 644 <file>"
    if "timeout" in combined:
        return "Increase LLM_TIMEOUT_SECONDS in .env, or reduce MAX_FILE_SIZE_MB to skip large files."
    if "unreadable" in combined or "encoding" in combined:
        return "Re-save the file as UTF-8. Run: `iconv -f <old> -t UTF-8 file > file.utf8`"
    if "large" in combined or "size" in combined:
        return f"File exceeds size limit. Split it into smaller modules or add to .gitignore."
    if "llm" in combined or "api" in combined:
        return "Check ANTHROPIC_API_KEY in .env. If key is valid, retry — may be a transient API error."
    return "Review the error message and re-run the scan after fixing the underlying issue."


def _fallbackSuggestion(issueType: str, description: str) -> str:
    """Ensure no suggestion is ever left blank."""
    it = issueType.lower()
    d  = description.lower()
    if "security" in it or "vulnerability" in it:
        return "Review the flagged code against OWASP Top 10 and apply the recommended security fix."
    if "syntax" in it:
        return "Fix the syntax error. Run a linter (pylint, eslint) to identify the exact problem."
    if "performance" in it:
        return "Profile and optimise the flagged code section to reduce resource usage."
    if "quality" in it or "best practice" in it:
        return "Refactor following clean-code principles and your team's style guide."
    return "Review and remediate the flagged issue before deploying to production."
